import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog


def open_controller():
    # 创建文件对话框

    result_label.config(text="Select the Controller Excel File")
    file_path = filedialog.askopenfilename(initialdir="/", title="Select the Controller Excel File")

    if file_path =='':
        result_label.config(text='Choose file please!')
        return
    controller_pd = pd.read_excel(file_path)

    # 讀取第二個Excel文件
    controller_template_excel = "./template/Controller Parameter Template.xlsx"
    template = pd.read_excel(controller_template_excel)

    controller_bike_model_template = template.iloc[:, [3, 5]]
    input_par = controller_pd.iloc[:, [2, 4]]

    bmt = controller_bike_model_template.set_index('Address')['Value'].to_dict()
    input = input_par.set_index('Address')['Value'].to_dict()

    controller_template_address = []

    for i in controller_bike_model_template['Address']:
        controller_template_address.append(i)

    print(controller_template_address)

    controller_input_value = []
    for i in controller_template_address:
        controller_input_value.append(input[i])
    print(controller_input_value)
    return controller_template_address, controller_input_value

def open_HMI():
    result_label.config(text="Select the HMI Excel File")
    file_path = filedialog.askopenfilename(initialdir="/", title="Select the HMI Excel File")

    if file_path =='':
        result_label.config(text='Choose file please!')
        return

    HMI_pd = pd.read_excel(file_path)
    # 讀取第二個Excel文件
    HMI_template_excel = "./template/HMI Parameter Template.xlsx"
    template = pd.read_excel(HMI_template_excel)

    HMI_bike_model_template = template.iloc[:, [3, 5]]
    input_par = HMI_pd.iloc[:, [2, 4]]

    bmt = HMI_bike_model_template.set_index('Address')['Value'].to_dict()
    input = input_par.set_index('Address')['Value'].to_dict()

    HMI_template_address = []

    for i in HMI_bike_model_template['Address']:
        HMI_template_address.append(i)

    print(HMI_template_address)

    HMI_input_value = []
    for i in HMI_template_address:
        HMI_input_value.append(input[i])
    print(HMI_input_value)

    return HMI_template_address, HMI_input_value


def to_excel():
    controller_template_address, controller_input_value = open_controller()
    HMI_template_address, HMI_input_value = open_HMI()
    # workbook = openpyxl.Workbook('0.xlsx')
    workbook = openpyxl.load_workbook('./template/SW_Bike_Model_Template.xlsx')
    # 选择默认的活动工作表
    # sheet = workbook.active
    #
    # sheet.title = 'Controller'
    sheet = workbook['Controller']

    sheet['D1'] = 'Address'
    sheet['F1'] = 'Value'

    for index, value in enumerate(controller_template_address, start=2):
        cell = sheet.cell(row=index, column=4, value=value)

    for index, value in enumerate(controller_input_value, start=2):
        cell = sheet.cell(row=index, column=6, value=value)

    # HMI_sheet = workbook.create_sheet("HMI")
    HMI_sheet = workbook['HMI']
    HMI_sheet['D1'] = 'Address'
    HMI_sheet['F1'] = 'Value'

    for index, value in enumerate(HMI_template_address, start=2):
        cell = HMI_sheet.cell(row=index, column=4, value=value)

    for index, value in enumerate(HMI_input_value, start=2):
        cell = HMI_sheet.cell(row=index, column=6, value=value)

    user_input = entry.get()
    if user_input == '':
        result_label.config(text='input name!')
        return
    workbook.save(user_input+".xlsx")

    # 保存工作簿到指定文件
    # workbook.save('example.xlsx')

    # 关闭工作簿
    workbook.close()
    result_label.config(text="done")



# 创建主窗口
root = tk.Tk()
root.title("Excel File Selection Example")
root.geometry("300x200")

show = tk.Label(root,text='Input Bike Model Name')
show.pack()


entry = tk.Entry(root)
entry.pack()

button_run = tk.Button(root, text='Create Bike Model', command=to_excel)
button_run.pack(pady=10)

result_label = tk.Label(root,text='')
result_label.pack()

# 启动 Tkinter 主循环
root.mainloop()
